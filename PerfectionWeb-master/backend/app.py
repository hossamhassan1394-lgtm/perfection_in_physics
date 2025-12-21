from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
import os
from dotenv import load_dotenv
from supabase import create_client, Client
import pandas as pd
from datetime import datetime
import traceback
import re
from dateutil import parser as date_parser
import logging
from logging.handlers import RotatingFileHandler
from collections import deque

# Load environment variables
load_dotenv()

app = Flask(__name__)
CORS(app, resources={
    r"/api/*": {
        "origins": ["*"],
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"],
        "supports_credentials": True
    }
})  # Enable CORS for Angular frontend

# Logging configuration
LOG_FILE = os.path.join(os.path.dirname(__file__), 'uploads.log')
logger = logging.getLogger('upload_logger')
logger.setLevel(logging.INFO)
handler = RotatingFileHandler(LOG_FILE, maxBytes=5 * 1024 * 1024, backupCount=3)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)
logger.propagate = False

# Configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Create uploads directory if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Supabase configuration
SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_KEY')

# Initialize supabase as None; will fail gracefully if not set
supabase: Client = None
if SUPABASE_URL and SUPABASE_KEY:
    try:
        supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
        logger.info("Supabase client initialized successfully")
    except Exception as e:
        logger.error(f"Failed to initialize Supabase: {str(e)}")
        supabase = None
else:
    logger.warning("SUPABASE_URL or SUPABASE_KEY not set - database operations will fail")

# Allowed groups (added 'online' option)
ALLOWED_GROUPS = ['cam1', 'maimi', 'cam2', 'west', 'station1', 'station2', 'station3', 'online']

# Allowed session numbers
ALLOWED_SESSIONS = list(range(1, 9))  # 1 to 8

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def normalize_phone(phone: str) -> str:
    """Normalize phone numbers to local format starting with '01' and 11 digits when possible.
    Examples:
      +201012345678 -> 01012345678
      201012345678 -> 01012345678
      01012345678 -> 01012345678
      1012345678   -> 01012345678
    """
    if not phone:
        return ''

    # Keep only digits
    cleaned = ''.join(ch for ch in phone if ch.isdigit())
    if cleaned.startswith('00'):
        cleaned = cleaned[2:]

    # If it has country code 20 (Egypt), convert to local starting with 0
    if cleaned.startswith('20') and len(cleaned) >= 11:
        candidate = '0' + cleaned[2:]
        if candidate.startswith('01') and len(candidate) == 11:
            return candidate

    # If it's already local 11-digit starting with 01
    if len(cleaned) == 11 and cleaned.startswith('01'):
        return cleaned

    # If it's 10 digits starting with 1 (missing leading zero)
    if len(cleaned) == 10 and cleaned.startswith('1'):
        return '0' + cleaned

    # As a last resort, if cleaned ends with 10 digits starting with '1', use that
    if len(cleaned) > 11 and cleaned[-10].isdigit():
        last10 = cleaned[-10:]
        if last10.startswith('1'):
            return '0' + last10

    return cleaned


def normalize_timestamp(value):
    """
    Normalize various timestamp inputs into a Postgres-friendly ISO string 'YYYY-MM-DD HH:MM:SS'.
    Handles:
      - Python datetimes
      - Strings with Arabic AM/PM markers (ص => AM, م => PM)
      - Common ambiguous formats (try dayfirst=True then False)
    Returns None when value is falsy.
    """
    if not value:
        return None

    # If already a datetime, return formatted string
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')

    s = str(value).strip()
    if not s:
        return None

    # Remove Unicode directionality marks and non-printable chars
    s = s.replace('\u200f', ' ').replace('\u200e', ' ')

    # Replace Arabic AM/PM markers with English equivalents
    # Arabic AM = 'ص' (U+0635), Arabic PM = 'م' (U+0645)
    s = s.replace('ص', ' AM').replace('م', ' PM')

    # Normalize whitespace
    s = re.sub(r'\s+', ' ', s).strip()

    # Try parsing with dayfirst True, then False
    for dayfirst in (True, False):
        try:
            dt = date_parser.parse(s, dayfirst=dayfirst)
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            continue

    # As a last resort return the original string (DB may still reject it)
    return s


def normalize_name(name: str) -> str:
    """Normalize a person name for matching: lower-case, collapse whitespace,
    remove common Arabic diacritics and Unicode combining marks to improve matching
    across uploads."""
    if not name:
        return ''
    try:
        import unicodedata
        s = str(name).strip().lower()
        # remove Arabic diacritics (harakat)
        s = re.sub(r'[\u064B-\u0652]', '', s)
        # normalize and remove combining marks
        nfkd = unicodedata.normalize('NFKD', s)
        s = ''.join(ch for ch in nfkd if not unicodedata.combining(ch))
        # collapse whitespace
        s = re.sub(r'\s+', ' ', s).strip()
        return s
    except Exception:
        try:
            return re.sub(r'\s+', ' ', str(name).strip().lower())
        except Exception:
            return str(name).strip().lower()


def format_start_time_arabic(value):
    """Format a datetime-like value into DD/MM/YYYY HH:MM:SS plus Arabic AM/PM marker (ص for AM, م for PM)."""
    # If value is missing/empty, show explicit placeholder for frontend
    if value is None or (isinstance(value, str) and not str(value).strip()):
        return 'No start time'
    try:
        if isinstance(value, datetime):
            dt = value
        else:
            dt = date_parser.parse(str(value))

        date_part = dt.strftime('%d/%m/%Y %H:%M:%S')
        arabic_marker = 'ص' if dt.hour < 12 else 'م'
        return f"{date_part} {arabic_marker}"
    except Exception:
        # If parsing fails, prefer returning the original non-empty string if available,
        # otherwise return the explicit placeholder so frontend shows "No start time".
        try:
            s = str(value)
            return s if s.strip() else 'No start time'
        except Exception:
            return 'No start time'

def parse_general_exam_sheet(file_path):
    """
    Parse general exam Excel sheet
    Expected columns: id, name, .Parent No, a, p, Q
    """
    try:
        # Try reading with header=0 first
        try:
            df = pd.read_excel(file_path, header=0)
        except AttributeError as ae:
            # Handle openpyxl ReadOnlyWorksheet lacking some attributes
            if 'defined_names' in str(ae) or 'ReadOnlyWorksheet' in str(ae):
                from openpyxl import load_workbook
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                data = list(ws.values)
                if not data:
                    raise
                header = [str(h).strip() if h is not None else '' for h in data[0]]
                rows = data[1:]
                df = pd.DataFrame(rows, columns=header)
            else:
                raise
        
        # Clean column names
        df.columns = df.columns.astype(str).str.strip()
        
        logger.info(f"General exam columns found: {list(df.columns)}")
        
        # Map columns by exact name or position
        col_list = list(df.columns)
        id_col = None
        name_col = None
        parent_col = None
        a_col = None
        p_col = None
        q_col = None
        
        # Try to find columns by name first (case-insensitive, stripped)
        for col in col_list:
            col_lower = str(col).lower().strip()
            
            if col_lower == 'id' and id_col is None:
                id_col = col
            elif col_lower == 'name' and name_col is None:
                name_col = col
            elif 'parent' in col_lower and parent_col is None:
                parent_col = col
            elif col_lower == 'a' and a_col is None:
                a_col = col
            elif col_lower == 'p' and p_col is None:
                p_col = col
            elif col_lower == 'q' and q_col is None:
                q_col = col
        
        # If not all found by name, try by position
        # Standard order: id, name, parent, a, p, q
        if not id_col and len(col_list) > 0:
            id_col = col_list[0]
        if not name_col and len(col_list) > 1:
            name_col = col_list[1]
        if not parent_col and len(col_list) > 2:
            parent_col = col_list[2]
        if not a_col and len(col_list) > 3:
            a_col = col_list[3]
        if not p_col and len(col_list) > 4:
            p_col = col_list[4]
        if not q_col and len(col_list) > 5:
            q_col = col_list[5]
        
        logger.info(f"Mapped columns - id:{id_col}, name:{name_col}, parent:{parent_col}, a:{a_col}, p:{p_col}, q:{q_col}")
        
        if not all([id_col, name_col, parent_col]):
            raise ValueError(f"Required columns (id, name, parent_no) not found. Columns: {list(df.columns)}")
        
        records = []
        for idx, row in df.iterrows():
            # Skip empty rows
            if pd.isna(row[id_col]) or str(row[id_col]).strip() == '':
                continue
            
            try:
                student_id = str(row[id_col]).strip()
                student_name = str(row[name_col]).strip() if not pd.isna(row[name_col]) else ''
                
                # Handle parent_no
                parent_no_val = row[parent_col]
                if pd.isna(parent_no_val):
                    parent_no_str = ''
                else:
                    try:
                        parent_no_str = str(int(float(parent_no_val)))
                    except Exception:
                        parent_no_str = str(parent_no_val).strip()
                
                # Validate required fields
                if not parent_no_str:
                    logger.warning(f"Row {idx+1}: Missing parent_no for student '{student_name}' (id='{student_id}')")
                    continue
                if not student_name:
                    logger.warning(f"Row {idx+1}: Missing student name for id='{student_id}'")
                    continue
                
                # Extract optional fields
                attendance = 0
                if a_col:
                    a_val = row[a_col]
                    if not pd.isna(a_val):
                        try:
                            attendance = 1 if (int(a_val) == 1 or str(a_val).strip() == '1') else 0
                        except (ValueError, TypeError):
                            attendance = 0
                
                payment = 0
                if p_col:
                    p_val = row[p_col]
                    if not pd.isna(p_val):
                        try:
                            payment = 1 if (int(p_val) == 1 or str(p_val).strip() == '1') else 0
                        except (ValueError, TypeError):
                            payment = 0
                
                quiz_mark = None
                if q_col:
                    q_val = row[q_col]
                    if not pd.isna(q_val):
                        try:
                            quiz_mark = float(q_val)
                        except (ValueError, TypeError):
                            quiz_mark = None
                
                record = {
                    'id': student_id,
                    'name': student_name,
                    'parent_no': parent_no_str,
                    'attendance': attendance,
                    'payment': payment,
                    'quiz_mark': quiz_mark
                }
                records.append(record)
                logger.info(f"Parsed row {idx+1}: {student_id} - {student_name}")
                
            except Exception as e:
                logger.warning(f"Error processing row {idx+1}: {str(e)}")
                continue
        
        logger.info(f"Parsed {len(records)} valid records from general exam sheet")
        return records
    except Exception as e:
        logger.error(f"Error parsing general exam sheet: {str(e)}")
        raise Exception(f"Error parsing general exam sheet: {str(e)}")

def parse_normal_lecture_sheet(file_path):
    """
    Parse normal lecture Excel sheet
    Columns: id, name, pokin, student no., Parent No., a, p, Q, time, s1
    """
    try:
        try:
            df = pd.read_excel(file_path, header=0)
        except AttributeError as ae:
            # Workaround for openpyxl returning ReadOnlyWorksheet without
            # `defined_names` when pandas/openpyxl open the file in
            # read-only mode. Fall back to loading with openpyxl directly
            # and build a DataFrame from the sheet values.
            if 'defined_names' in str(ae):
                from openpyxl import load_workbook
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                data = list(ws.values)
                if not data:
                    raise
                header = [str(h).strip() if h is not None else '' for h in data[0]]
                rows = data[1:]
                df = pd.DataFrame(rows, columns=header)
            else:
                raise
        
        # Clean column names
        df.columns = df.columns.astype(str).str.strip()
        
        # Find columns
        id_col = None
        name_col = None
        pokin_col = None
        student_no_col = None
        parent_col = None
        a_col = None
        p_col = None
        q_col = None
        time_col = None
        s1_col = None
        
        for col in df.columns:
            col_lower = str(col).lower().strip()
            if 'id' in col_lower and id_col is None and col_lower != 'parent' and 'student' not in col_lower:
                id_col = col
            elif 'name' in col_lower and name_col is None:
                name_col = col
            elif 'pokin' in col_lower and pokin_col is None:
                pokin_col = col
            elif 'student' in col_lower and 'no' in col_lower and student_no_col is None:
                student_no_col = col
            elif 'parent' in col_lower and 'no' in col_lower and parent_col is None:
                parent_col = col
            elif col_lower == 'a' and a_col is None:
                a_col = col
            elif col_lower == 'p' and p_col is None:
                p_col = col
            elif col_lower == 'q' and q_col is None:
                q_col = col
            elif 'time' in col_lower and time_col is None:
                time_col = col
            elif col_lower == 's1' and s1_col is None:
                s1_col = col
        
        if not all([id_col, name_col, parent_col]):
            raise ValueError(f"Required columns not found in normal lecture sheet. Found: {list(df.columns)}")
        
        records = []
        for _, row in df.iterrows():
            if pd.isna(row[id_col]) or str(row[id_col]).strip() == '':
                continue
            
            try:
                # Parse time if available
                start_time = None
                if time_col and not pd.isna(row[time_col]):
                    try:
                        time_val = row[time_col]
                        if isinstance(time_val, datetime):
                            start_time = time_val.strftime('%Y-%m-%d %H:%M:%S')
                        elif isinstance(time_val, pd.Timestamp):
                            start_time = time_val.strftime('%Y-%m-%d %H:%M:%S')
                        else:
                            start_time = str(time_val).strip()
                            # Only set if not empty
                            if not start_time:
                                start_time = None
                    except Exception as time_error:
                        logger.warning(f"Warning parsing time for {row[id_col]}: {str(time_error)}")
                        start_time = None
                
                # Handle parent_no conversion (CRITICAL - must not be empty if provided)
                parent_no_val = row[parent_col]
                parent_no_str = ''
                if not pd.isna(parent_no_val) and str(parent_no_val).strip():
                    try:
                        # Try to parse as number first
                        parent_no_str = str(int(float(parent_no_val)))
                    except Exception:
                        parent_no_str = str(parent_no_val).strip()
                
                # Handle student_no conversion
                student_no_str = None
                if student_no_col and not pd.isna(row[student_no_col]) and str(row[student_no_col]).strip():
                    try:
                        student_no_str = str(int(float(row[student_no_col])))
                    except Exception:
                        student_no_str = str(row[student_no_col]).strip()
                
                # Parse pokin
                pokin_val = None
                if pokin_col and not pd.isna(row[pokin_col]):
                    try:
                        pokin_val = float(row[pokin_col])
                    except Exception:
                        pass
                
                # Parse payment
                payment_val = None
                if p_col and not pd.isna(row[p_col]):
                    try:
                        payment_val = float(row[p_col])
                    except Exception:
                        pass
                
                # Parse quiz mark
                quiz_val = None
                if q_col and not pd.isna(row[q_col]):
                    try:
                        quiz_val = float(row[q_col])
                    except Exception:
                        pass
                
                record = {
                    'id': str(row[id_col]).strip(),
                    'name': str(row[name_col]).strip() if not pd.isna(row[name_col]) else '',
                    'pokin': pokin_val,
                    'student_no': student_no_str,
                    'parent_no': parent_no_str,
                    'attendance': 1 if a_col and not pd.isna(row[a_col]) and (row[a_col] == 1 or str(row[a_col]).strip() == '1') else 0,
                    'payment': payment_val,
                    'quiz_mark': quiz_val,
                    'start_time': start_time,
                    'homework_status': 0  # Default to completed
                }
                
                # Parse s1 (homework status)
                # null/empty = completed (0), 1 = no hw, 2 = not completed, 3 = cheated
                if s1_col and not pd.isna(row[s1_col]) and str(row[s1_col]).strip():
                    try:
                        s1_value = int(float(row[s1_col]))
                        record['homework_status'] = s1_value
                    except:
                        record['homework_status'] = 0
                
                records.append(record)
            except Exception as e:
                # Skip rows with errors but continue processing
                logger.warning(f"Error processing row {row.get(id_col, 'unknown')}: {str(e)}")
                continue
        
        return records
    except Exception as e:
        raise Exception(f"Error parsing normal lecture sheet: {str(e)}")

def create_or_update_parent(parent_no, student_name=None):
    """
    Create or update parent account in parents table
    Default password is '123456' and needs_password_reset is True
    """
    if not parent_no or parent_no.strip() == '':
        return

    parent_no_norm = normalize_phone(parent_no)
    
    try:
        # Check if parent exists
        existing = supabase.table('parents').select('*').eq('phone_number', parent_no_norm).execute()
        if existing.data and len(existing.data) > 0:
            # Parent exists, no need to update
            return
        else:
            # Create new parent account
            parent_data = {
                'phone_number': parent_no_norm,
                'password_hash': '123456',  # Default password (in production, hash this)
                'needs_password_reset': True,
                'name': student_name or f'Parent {parent_no_norm}'
            }
            try:
                res = supabase.table('parents').insert(parent_data).execute()
                if getattr(res, 'error', None):
                    logger.warning(f"Could not create parent {parent_no_norm}: {res.error}")
            except Exception as e:
                logger.exception(f"Exception creating parent {parent_no_norm}: {str(e)}")
    except Exception as e:
        # Silently fail parent creation - don't block the main upload
        logger.warning(f"Could not create parent account for {parent_no}: {str(e)}")

def update_database(records, session_number, quiz_mark, finish_time, group, is_general_exam, lecture_name='', exam_name='', has_exam_grade=True, has_payment=True, has_time=True):
    """
    Update database with parsed records using UPSERT logic
    Works with existing constraint: UNIQUE (student_name, session_number, parent_no)
    """
    try:
        updated_count = 0
        errors = []
        
        # Collect unique parent_no values from all records to pre-fetch existing records
        unique_parent_nos = set()
        for rec in records:
            parent_no_raw = rec.get('parent_no', '') or ''
            parent_no = normalize_phone(parent_no_raw) or ''
            if parent_no:
                unique_parent_nos.add(parent_no)
        
        # Pre-fetch all existing records for these parent numbers (cache lookup)
        existing_by_parent = {}
        if unique_parent_nos:
            try:
                for parent_no in unique_parent_nos:
                    existing = supabase.table('session_records').select('student_id', 'student_name').eq('parent_no', parent_no).execute()
                    if getattr(existing, 'data', None):
                        existing_by_parent[parent_no] = existing.data
                    else:
                        existing_by_parent[parent_no] = []
            except Exception as e:
                logger.warning("Error pre-fetching existing records: %s", str(e))
        
        # Batch lists for insert and update operations
        inserts_to_do = []
        updates_to_do = []  # (old_id, update_data)
        
        for record in records:
            try:
                student_id = record.get('id', '').strip()
                student_name = record.get('name', '').strip() or 'Unknown'
                parent_no_raw = record.get('parent_no', '') or ''
                parent_no = normalize_phone(parent_no_raw) or ''

                # Validate required fields
                if not parent_no:
                    msg = f"Missing parent_no for student '{student_name}' (raw='{parent_no_raw}')"
                    logger.warning(msg)
                    errors.append(msg)
                    continue
                
                if not student_name or student_name == 'Unknown':
                    msg = f"Missing student_name for id '{student_id}'"
                    logger.warning(msg)
                    errors.append(msg)
                    continue
                
                # Prepare data for database
                db_data = {
                    'student_id': student_id or f'student_{updated_count}',
                    'student_name': student_name,
                    'parent_no': parent_no,
                    'session_number': session_number,
                    'group_name': group,
                    'is_general_exam': bool(is_general_exam),  # CRITICAL: Ensure boolean
                    'attendance': int(record.get('attendance', 0)) if record.get('attendance') else 0,
                    'payment': float(record.get('payment', 0)) if record.get('payment') else 0,
                }
                
                # Add lecture/exam name
                if is_general_exam and exam_name:
                    db_data['exam_name'] = exam_name
                elif not is_general_exam and lecture_name:
                    db_data['lecture_name'] = lecture_name
                
                # Add admin quiz mark
                if quiz_mark is not None:
                    db_data['admin_quiz_mark'] = float(quiz_mark)
                
                # Add optional fields
                if record.get('quiz_mark') is not None:
                    db_data['quiz_mark'] = float(record.get('quiz_mark'))
                
                if finish_time:
                    try:
                        normalized_finish = normalize_timestamp(finish_time)
                        db_data['finish_time'] = normalized_finish if normalized_finish else finish_time
                    except Exception:
                        db_data['finish_time'] = finish_time

                if record.get('start_time'):
                    # Normalize start_time from the parsed record (handles Arabic AM/PM, etc.)
                    try:
                        normalized_start = normalize_timestamp(record.get('start_time'))
                        db_data['start_time'] = normalized_start if normalized_start else record.get('start_time')
                    except Exception:
                        db_data['start_time'] = record.get('start_time')
                
                if record.get('homework_status') is not None:
                    db_data['homework_status'] = int(record.get('homework_status'))
                
                if record.get('pokin'):
                    db_data['pokin'] = float(record.get('pokin'))
                
                if record.get('student_no'):
                    db_data['student_no'] = str(record.get('student_no')).strip()
                
                # Try to insert record; handle Supabase response errors (client may not raise)
                try:
                    insert_res = supabase.table('session_records').insert(db_data).execute()
                    insert_error = getattr(insert_res, 'error', None)
                    if not insert_error:
                        updated_count += 1
                        logger.info(f"Inserted record for {student_id} (session {session_number}, group {group})")
                    else:
                        # Log full error dict + payload for diagnostics
                        try:
                            logger.error("Insert error for %s: %s -- payload: %s", student_id, insert_error, db_data)
                        except Exception:
                            logger.error("Insert error for %s: %s", student_id, str(insert_error))

                        # Normalize error message text for checks and append to errors
                        error_msg = None
                        try:
                            error_msg = insert_error.get('message') if isinstance(insert_error, dict) else str(insert_error)
                        except Exception:
                            error_msg = str(insert_error)

                        # If the error is caused by a missing column in the schema cache
                        # (e.g. admin_quiz_mark), try removing that column and retrying.
                        if isinstance(error_msg, str) and ("Could not find the 'admin_quiz_mark'" in error_msg or 'PGRST204' in str(insert_error)):
                            try:
                                reduced_payload = dict(db_data)
                                if 'admin_quiz_mark' in reduced_payload:
                                    reduced_payload.pop('admin_quiz_mark')
                                # attempt insert without the offending column
                                retry_res = supabase.table('session_records').insert(reduced_payload).execute()
                                retry_error = getattr(retry_res, 'error', None)
                                if not retry_error:
                                    updated_count += 1
                                    logger.info(f"Inserted record for %s after removing admin_quiz_mark (session %s, group %s)", student_id, session_number, group)
                                    # remove last recorded error if present
                                    if errors:
                                        errors.pop()
                                    continue
                                else:
                                    logger.error("Retry insert error for %s: %s -- reduced payload: %s", student_id, retry_error, reduced_payload)
                            except Exception as retry_exc:
                                logger.exception("Retry insert exception for %s: %s", student_id, str(retry_exc))

                        # If duplicate key (unique constraint) - try updating the existing row instead of failing
                        if isinstance(error_msg, str) and ('23505' in error_msg or 'duplicate' in error_msg.lower() or 'unique' in error_msg.lower() or 'session_records_student_name_session_number_parent_no_key' in error_msg):
                            try:
                                # Preferred: update by the desired unique key (student_name, parent_no)
                                update_res = supabase.table('session_records').update(db_data).eq('student_name', student_name).eq('parent_no', parent_no).execute()
                                update_error = getattr(update_res, 'error', None)
                                if not update_error:
                                    updated_count += 1
                                    logger.info("Updated existing record by student_name+parent_no for %s", student_id)
                                    if errors:
                                        errors.pop()
                                    continue
                            except Exception:
                                logger.exception("Exception while attempting update by student_name+parent_no for %s", student_id)

                            # Fallback: try targeted update by student_id + session/group
                            try:
                                fut_update = supabase.table('session_records').update(db_data).eq('student_id', student_id).eq('session_number', session_number).eq('group_name', group).eq('is_general_exam', is_general_exam).execute()
                                fut_err = getattr(fut_update, 'error', None)
                                if not fut_err:
                                    updated_count += 1
                                    logger.info("Updated duplicate record for %s via fallback keys", student_id)
                                    if errors:
                                        errors.pop()
                                    continue
                                else:
                                    logger.error("Fallback update also failed for %s: %s", student_id, fut_err)
                            except Exception:
                                logger.exception("Fallback update exception for %s", student_id)

                        # Record detailed error for response
                        errors.append(f"Row {student_id}: {error_msg} | payload: {db_data}")

                        # As a last attempt, try to find an existing record by name/session/group and update student_id if it differs
                        try:
                            existing = supabase.table('session_records').select('student_id').eq('student_name', student_name).eq('session_number', session_number).eq('group_name', group).eq('is_general_exam', is_general_exam).limit(1).execute()
                            if getattr(existing, 'data', None) and len(existing.data) > 0:
                                old_id = existing.data[0].get('student_id')
                                if old_id and old_id != student_id:
                                    logger.info("Found same person '%s' with old ID '%s', updating to new ID '%s'", student_name, old_id, student_id)
                                    update_data = dict(db_data)
                                    update_data['student_id'] = student_id
                                    try:
                                        update_by_name = supabase.table('session_records').update(update_data).eq('student_id', old_id).eq('session_number', session_number).eq('group_name', group).eq('is_general_exam', is_general_exam).execute()
                                        update_name_error = getattr(update_by_name, 'error', None)
                                        if not update_name_error:
                                            updated_count += 1
                                            logger.info("Updated student ID from '%s' to '%s' for '%s'", old_id, student_id, student_name)
                                            if errors:
                                                errors.pop()
                                        else:
                                            name_err = update_name_error.get('message') if isinstance(update_name_error, dict) else str(update_name_error)
                                            logger.error("Name-based update failed for %s: %s -- payload: %s", student_name, name_err, db_data)
                                    except Exception:
                                        logger.exception("Name-based update exception for %s", student_name)
                        except Exception as name_err:
                            logger.exception("Name-based lookup/update exception for %s: %s", student_name, str(name_err))
                except Exception as e:
                    # Exception from Supabase client call
                    err_text = str(e)
                    errors.append(f"Row {student_id}: {err_text}")
                    logger.exception(f"Insert exception for {student_id}: {err_text}")
                    
            except Exception as e:
                errors.append(str(e))
        
        logger.info(f"Upload summary: {updated_count}/{len(records)} records uploaded, {len(errors)} errors")
        return updated_count, errors
        
    except Exception as e:
        logger.exception(f"✗ Critical error in update_database: {str(e)}")
        raise Exception(f"Error updating database: {str(e)}")

@app.route('/', methods=['GET'])
def root():
    """Root endpoint - returns API info"""
    return jsonify({'message': 'Perfection Physics Backend API', 'version': '1.0', 'status': 'running'}), 200

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({'status': 'ok', 'message': 'Flask backend is running'}), 200
        
@app.route('/api/upload-excel', methods=['POST'])
def upload_excel():
    """
    Upload and process Excel file
    Expected form data:
    - file: Excel file
    - session_number: 1-8
    - quiz_mark: number (for general exam)
    - finish_time: datetime string
    - group: cam1, maimi, cam2, west, station1, station2, station3, online
    - is_general_exam: true/false
    - lecture_name: string (for normal lectures)
    - exam_name: string (for general exams)
    - has_exam_grade: true/false (show exam grade in parent dashboard)
    - has_payment: true/false (show payment in parent dashboard)
    - has_time: true/false (show finish time in parent dashboard)
    """
    try:
        # Check if Supabase is initialized
        if not supabase:
            return jsonify({'error': 'Database not configured. Please contact administrator.'}), 500
        
        # Validate required fields
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Invalid file type. Only .xlsx and .xls files are allowed'}), 400
        
        # Get form data
        session_number = request.form.get('session_number')
        quiz_mark = request.form.get('quiz_mark')
        finish_time = request.form.get('finish_time')
        group = request.form.get('group')
        is_general_exam = request.form.get('is_general_exam', 'false').lower() == 'true'
        lecture_name = request.form.get('lecture_name', '').strip()
        # Optional: allow passing a lecture unique key which maps to a lecture_name
        lecture_key = request.form.get('lecture_key', '').strip()
        exam_name = request.form.get('exam_name', '').strip()
        has_exam_grade = request.form.get('has_exam_grade', 'true').lower() == 'true'
        has_payment = request.form.get('has_payment', 'true').lower() == 'true'
        has_time = request.form.get('has_time', 'true').lower() == 'true'

        # Validate session number
        try:
            session_number = int(session_number)
            if session_number not in ALLOWED_SESSIONS:
                return jsonify({'error': f'Session number must be between 1 and 8'}), 400
        except (ValueError, TypeError):
            return jsonify({'error': 'Invalid session number'}), 400
        
        # Validate quiz mark (required for general exam)
        if is_general_exam:
            try:
                quiz_mark = float(quiz_mark) if quiz_mark else None
            except (ValueError, TypeError):
                return jsonify({'error': 'Invalid quiz mark'}), 400
        else:
            quiz_mark = float(quiz_mark) if quiz_mark else None
        
        # Validate group
        if not group or group not in ALLOWED_GROUPS:
            return jsonify({'error': f'Invalid group. Must be one of: {", ".join(ALLOWED_GROUPS)}'}), 400
        
        # Save file
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        try:
            # Parse Excel file based on type
            if is_general_exam:
                records = parse_general_exam_sheet(file_path)
            else:
                records = parse_normal_lecture_sheet(file_path)
            
            if not records:
                return jsonify({'error': 'No records found in Excel file'}), 400
            
            # If lecture_name not provided but lecture_key exists, lookup lecture_name from `lectures` table
            if not lecture_name and lecture_key:
                try:
                    lookup = supabase.table('lectures').select('lecture_name').eq('unique_key', lecture_key).limit(1).execute()
                    if lookup.data and len(lookup.data) > 0:
                        lecture_name = lookup.data[0].get('lecture_name', '') or lecture_name
                except Exception as e:
                    logger.warning(f"Could not resolve lecture_key {lecture_key}: {str(e)}")

            # Update database
            updated_count, errors = update_database(
                records, 
                session_number, 
                quiz_mark, 
                finish_time, 
                group, 
                is_general_exam,
                lecture_name,
                exam_name,
                has_exam_grade,
                has_payment,
                has_time
            )
            
            # Clean up uploaded file
            os.remove(file_path)
            
            response = {
                'success': True,
                'message': f'Successfully processed {updated_count} records',
                'updated_count': updated_count,
                'total_records': len(records),
                'partial': False
            }

            if errors:
                response['errors'] = errors
                response['error_count'] = len(errors)
                # If some records succeeded but some failed, mark as partial
                if updated_count > 0:
                    response['partial'] = True
                    response['message'] = f'Processed {updated_count}/{len(records)} records with {len(errors)} errors'
                    response['success'] = True
                else:
                    # All records failed
                    response['partial'] = False
                    response['success'] = False
                    response['message'] = f'All records failed: {len(errors)} errors'

            return jsonify(response), 200
            
        except Exception as e:
            # Clean up file on error
            if os.path.exists(file_path):
                os.remove(file_path)
            return jsonify({'error': f'Error processing file: {str(e)}', 'traceback': traceback.format_exc()}), 500
        
    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}', 'traceback': traceback.format_exc()}), 500

@app.route('/api/groups', methods=['GET'])
def get_groups():
    """Get list of available groups"""
    return jsonify({'groups': ALLOWED_GROUPS}), 200

@app.route('/api/sessions', methods=['GET'])
def get_sessions():
    """Get list of available session numbers"""
    return jsonify({'sessions': ALLOWED_SESSIONS}), 200


@app.route('/api/parent/students', methods=['GET'])
def get_parent_students():
    """Return aggregated student list for a parent, grouped by parent_no + student_name"""
    phone = request.args.get('phone_number')
    if not phone:
        return jsonify({'error': 'phone_number query parameter required'}), 400
    phone = normalize_phone(phone)

    try:
        result = supabase.table('session_records').select('*').eq('parent_no', phone).execute()
        records = result.data or []
        # Debug logging: report counts and sample flags
        try:
            total_records = len(records)
            ge_count = sum(1 for rr in records if (rr.get('is_general_exam') in (True, 'true', 1) or str(rr.get('is_general_exam')).lower() == 'true'))
            logger.info("Parent sessions fetch for %s: %d records, %d general_exam flags", phone, total_records, ge_count)
            # Log first 5 records simplified
            for sample in records[:5]:
                logger.info("Sample record: id=%s student_id=%s student_name=%s is_general_exam=%s session_number=%s parent_no=%s", sample.get('id'), sample.get('student_id'), sample.get('student_name'), sample.get('is_general_exam'), sample.get('session_number'), sample.get('parent_no'))
        except Exception:
            logger.exception("Error logging parent session debug info")

        # Group by student_name (matches database UNIQUE constraint)
        students_map = {}
        for r in records:
            student_name = (r.get('student_name') or '').strip()
            if not student_name:
                continue
            
            # Use student_name as key (stable across uploads)
            if student_name not in students_map:
                students_map[student_name] = {
                    'name': student_name,
                    'parent_no': phone,
                    'ids': set(),
                    'grade': '',
                    'attendance_count': 0,
                    'records_count': 0,
                    'payments_sum': 0.0,
                    'quiz_sum': 0.0,
                    'quiz_count': 0
                }
            
            entry = students_map[student_name]
            
            # Track all student_ids (for reference)
            student_id = r.get('student_id')
            if student_id:
                entry['ids'].add(student_id)
            
            entry['records_count'] += 1
            entry['attendance_count'] += int(r.get('attendance', 0))
            entry['payments_sum'] += float(r.get('payment', 0))
            
            quiz_mark = r.get('quiz_mark')
            if quiz_mark is not None:
                entry['quiz_sum'] += float(quiz_mark)
                entry['quiz_count'] += 1

        students = []
        for name, v in students_map.items():
            attendance_pct = round((v['attendance_count'] / v['records_count']) * 100) if v['records_count'] > 0 else 0
            total_expected = v['records_count'] * 140
            quizzes_avg = round((v['quiz_sum'] / v['quiz_count']), 2) if v['quiz_count'] > 0 else 0

            students.append({
                'id': v['parent_no'],  # Use parent_no as stable ID
                'name': v['name'],
                'grade': v.get('grade', ''),
                'attendance': attendance_pct,
                'payments': {'paid': v['payments_sum'], 'total': total_expected},
                'quizzes': {'average': quizzes_avg, 'total': v['quiz_count']}
            })

        return jsonify({'students': students}), 200
    except Exception as e:
        logger.exception(f"Error fetching students: {str(e)}")
        return jsonify({'error': f'Error fetching students: {str(e)}'}), 500
    
    
@app.route('/api/parent/sessions', methods=['GET'])
def get_parent_sessions():
    """Return session records for a parent with proper boolean handling"""
    phone = request.args.get('phone_number')
    if not phone:
        return jsonify({'error': 'phone_number query parameter required'}), 400
    phone = normalize_phone(phone)

    try:
        query = supabase.table('session_records').select('*').eq('parent_no', phone)
        result = query.execute()
        records = result.data or []

        sessions = []
        for r in records:
            has_exam_grade = r.get('has_exam_grade', True)
            has_payment = r.get('has_payment', True)
            has_time = r.get('has_time', True)
            
            # CRITICAL FIX: Properly handle is_general_exam boolean
            is_general_exam_raw = r.get('is_general_exam')
            is_general_exam = False
            
            # Handle all possible true values
            if is_general_exam_raw is True:
                is_general_exam = True
            elif isinstance(is_general_exam_raw, str) and is_general_exam_raw.lower() == 'true':
                is_general_exam = True
            elif isinstance(is_general_exam_raw, int) and is_general_exam_raw == 1:
                is_general_exam = True
            
            formatted_start = format_start_time_arabic(r.get('start_time'))

            session = {
                'id': r.get('id') or r.get('student_no') or r.get('student_id'),
                'chapter': r.get('session_number'),
                'name': r.get('lecture_name') or r.get('exam_name') or f"Session {r.get('session_number')}",
                'lectureName': r.get('lecture_name') or r.get('exam_name'),
                'date': r.get('finish_time') or '',
                # include created_at so frontend can order by upload time
                'created_at': r.get('created_at') or r.get('createdAt') or r.get('created at'),
                'is_general_exam': r.get('is_general_exam', False),
                'isGeneralExam': r.get('is_general_exam', False),
                'startTime': formatted_start,
                'start_time': formatted_start,
                'attendance': 'attended' if int(r.get('attendance') or 0) == 1 else 'missed',
                'homeworkStatus': 'completed' if (r.get('homework_status') in (0, None)) else 'pending',
                'is_general_exam': is_general_exam,
                'isGeneralExam': is_general_exam
            }
            
            if has_exam_grade:
                quiz_mark = int(r.get('quiz_mark') or 0)
                admin_quiz_mark = r.get('admin_quiz_mark')
                
                session['quizCorrect'] = quiz_mark
                
                if admin_quiz_mark is not None:
                    session['adminQuizMark'] = int(admin_quiz_mark)
                    session['quizTotal'] = int(admin_quiz_mark)
                else:
                    session['quizTotal'] = 15
            
            if has_payment:
                session['payment'] = float(r.get('payment') or 0)
            
            if has_time:
                session['endTime'] = r.get('finish_time') or ''
            
            sessions.append(session)

        # Prefer ordering by upload/creation time (newest first). If created_at
        # is not present, fall back to ordering by chapter/session number.
        sessions = sorted(
            sessions,
            key=lambda s: ((s.get('created_at') or ''), (s.get('chapter') or 0)),
            reverse=True
        )
        return jsonify({'sessions': sessions}), 200
        
    except Exception as e:
        logger.exception(f"Error fetching sessions: {str(e)}")
        return jsonify({'error': f'Error fetching sessions: {str(e)}'}), 500
    
    
@app.route('/api/students', methods=['GET'])
def get_all_students():
    """Return aggregated student list across all parents (for admin)"""
    try:
        result = supabase.table('session_records').select('*').execute()
        records = result.data or []

        students_map = {}
        for r in records:
            sid = r.get('student_id')
            if not sid:
                continue
            entry = students_map.setdefault(sid, {
                'id': sid,
                'name': r.get('student_name') or '',
                'grade': '',
                'attendance_count': 0,
                'records_count': 0,
                'payments_sum': 0.0,
                'quiz_sum': 0.0,
                'quiz_count': 0
            })

            entry['records_count'] += 1
            try:
                entry['attendance_count'] += int(r.get('attendance') or 0)
            except:
                pass
            try:
                entry['payments_sum'] += float(r.get('payment') or 0)
            except:
                pass
            try:
                if r.get('quiz_mark') is not None:
                    entry['quiz_sum'] += float(r.get('quiz_mark'))
                    entry['quiz_count'] += 1
            except:
                pass

        students = []
        for sid, v in students_map.items():
            attendance_pct = 0
            if v['records_count'] > 0:
                attendance_pct = round((v['attendance_count'] / v['records_count']) * 100)

            total_expected = v['records_count'] * 140
            quizzes_avg = round((v['quiz_sum'] / v['quiz_count']), 2) if v['quiz_count'] > 0 else 0

            students.append({
                'id': v['id'],
                'name': v['name'],
                'grade': v.get('grade', ''),
                'attendance': attendance_pct,
                'payments': { 'paid': v['payments_sum'], 'total': total_expected },
                'quizzes': { 'average': quizzes_avg, 'total': v['quiz_count'] }
            })

        return jsonify({'students': students}), 200
    except Exception as e:
        return jsonify({'error': f'Error fetching all students: {str(e)}', 'traceback': traceback.format_exc()}), 500

@app.route('/api/auth/login', methods=['POST'])
def login():
    """
    Parent login endpoint
    Expected JSON:
    {
        "phone_number": "01234567890",
        "password": "123456"
    }
    """
    try:
        data = request.get_json()
        phone_number = data.get('phone_number', '').strip()
        password = data.get('password', '').strip()
        
        if not phone_number or not password:
            return jsonify({'success': False, 'message': 'Phone number and password are required'}), 400
        
        # Normalize phone and find parent by phone number
        phone_number = normalize_phone(phone_number)
        result = supabase.table('parents').select('*').eq('phone_number', phone_number).execute()
        
        if not result.data or len(result.data) == 0:
            # Parent doesn't exist, create a new one with default password
            logger.info(f"Creating new parent account for {phone_number}")
            try:
                parent_data = {
                    'phone_number': phone_number,
                    'password_hash': password,
                    'needs_password_reset': True,
                    'name': f'Parent {phone_number}'
                }
                create_result = supabase.table('parents').insert(parent_data).execute()
                if create_result and create_result.data:
                    parent = create_result.data[0]
                    logger.info(f"Parent account created for {phone_number}")
                else:
                    return jsonify({'success': False, 'message': 'Failed to create parent account'}), 500
            except Exception as create_error:
                logger.exception(f"Error creating parent: {str(create_error)}")
                return jsonify({'success': False, 'message': f'Error creating account: {str(create_error)}'}), 500
        else:
            parent = result.data[0]
            
            # Check password (in production, use proper password hashing)
            if parent['password_hash'] != password:
                return jsonify({'success': False, 'message': 'Invalid phone number or password'}), 401
        
        # Update last login
        try:
            supabase.table('parents').update({'last_login': datetime.now().isoformat()}).eq('phone_number', phone_number).execute()
        except:
            pass  # Don't fail login if update fails
        
        # Return user data
        return jsonify({
            'success': True,
            'user': {
                'phone_number': parent['phone_number'],
                'name': parent.get('name', ''),
                'needs_password_reset': parent.get('needs_password_reset', True)
            },
            'needs_password_reset': parent.get('needs_password_reset', True)
        }), 200
        
    except Exception as e:
        logger.exception(f"Login error: {str(e)}")
        return jsonify({'success': False, 'message': f'Login error: {str(e)}'}), 500


@app.route('/api/auth/reset-password', methods=['POST'])
def reset_password():
    """
    Reset password for first-time login.
    Expected JSON:
    {
        "phone_number": "01234567890",
        "new_password": "newpassword123"
    }
    """
    try:
        data = request.get_json() or {}
        phone_number = (data.get('phone_number') or '').strip()
        new_password = (data.get('new_password') or '').strip()

        if not phone_number or not new_password:
            return jsonify({'success': False, 'message': 'Phone number and new password are required'}), 400

        if len(new_password) < 6:
            return jsonify({'success': False, 'message': 'Password must be at least 6 characters long'}), 400

        phone_number = normalize_phone(phone_number)

        # Find parent
        result = supabase.table('parents').select('*').eq('phone_number', phone_number).execute()
        if not result.data or len(result.data) == 0:
            return jsonify({'success': False, 'message': 'Parent not found'}), 404

        try:
            update_result = supabase.table('parents').update({
                'password_hash': new_password,
                'needs_password_reset': False
            }).eq('phone_number', phone_number).execute()

            if update_result and update_result.data:
                return jsonify({'success': True, 'message': 'Password updated successfully'}), 200
            else:
                return jsonify({'success': False, 'message': 'Failed to update password'}), 500
        except Exception as e:
            return jsonify({'success': False, 'message': f'Error updating password: {str(e)}'}), 500

    except Exception as e:
        return jsonify({'success': False, 'message': f'Reset password error: {str(e)}'}), 500

@app.route('/api/auth/change-password', methods=['POST'])
def change_password():
    """
    Change password endpoint for parents
    Expected JSON:
    {
        "phone_number": "01234567890",
        "current_password": "oldpassword",
        "new_password": "newpassword123"
    }
    """
    try:
        data = request.get_json()
        phone_number = data.get('phone_number', '').strip()
        current_password = data.get('current_password', '').strip()
        new_password = data.get('new_password', '').strip()
        
        if not phone_number or not current_password or not new_password:
            return jsonify({'success': False, 'message': 'Phone number, current password, and new password are required'}), 400
        
        if len(new_password) < 6:
            return jsonify({'success': False, 'message': 'Password must be at least 6 characters long'}), 400
        
        # Normalize phone and find parent
        phone_number = normalize_phone(phone_number)
        result = supabase.table('parents').select('*').eq('phone_number', phone_number).execute()
        
        if not result.data or len(result.data) == 0:
            return jsonify({'success': False, 'message': 'Parent not found'}), 404
        
        parent = result.data[0]
        
        # Check if password_hash column exists and has a value
        if 'password_hash' not in parent or not parent['password_hash']:
            # For new parents without password set, allow setting one without verification
            stored_password = parent.get('password') or parent.get('password_hash')
        else:
            stored_password = parent.get('password_hash')
        
        # Verify current password (in production, use proper password hashing)
        if stored_password and stored_password != current_password:
            return jsonify({'success': False, 'message': 'Current password is incorrect'}), 401
        
        # Update to new password - try password_hash first, then password
        try:
            # First, try to update with password_hash column
            try:
                update_result = supabase.table('parents').update({
                    'password_hash': new_password
                }).eq('phone_number', phone_number).execute()
                
                # Check if update actually worked by verifying the data was updated
                if update_result and update_result.data:
                    return jsonify({
                        'success': True,
                        'message': 'Password changed successfully'
                    }), 200
            except Exception as hash_error:
                # If password_hash fails, try password column instead
                logger.warning(f"password_hash update failed, trying password column: {str(hash_error)}")
                try:
                    update_result = supabase.table('parents').update({
                        'password': new_password
                    }).eq('phone_number', phone_number).execute()
                    
                    if update_result and update_result.data:
                        return jsonify({
                            'success': True,
                            'message': 'Password changed successfully'
                        }), 200
                except Exception as password_error:
                    logger.exception(f"password update failed: {str(password_error)}")
                    raise password_error
            
            # If we get here, something went wrong
            return jsonify({
                'success': False,
                'message': 'Failed to update password in database'
            }), 500
        except Exception as update_error:
            error_msg = str(update_error)
            logger.exception(f"Database update error: {error_msg}")
            return jsonify({'success': False, 'message': f'Database error: {error_msg}'}), 500
        
    except Exception as e:
        logger.exception(f"Change password error: {str(e)}")
        return jsonify({'success': False, 'message': f'Error changing password: {str(e)}'}), 500
        
        # Update password (in production, hash the password)
        supabase.table('parents').update({
            'password_hash': new_password,
            'needs_password_reset': False
        }).eq('phone_number', phone_number).execute()
        
        return jsonify({
            'success': True,
            'message': 'Password updated successfully'
        }), 200
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error updating password: {str(e)}'}), 500

@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    """Admin login endpoint. Expects JSON: { username, password }"""
    print("\n" + "="*80)
    print("🔐 ADMIN LOGIN ATTEMPT STARTED")
    print("="*80)
    
    try:
        data = request.get_json() or {}
        username = (data.get('username') or '').strip()
        password = (data.get('password') or '').strip()

        print(f"📝 Received data: {data}")
        print(f"👤 Username: '{username}'")
        print(f"🔑 Password: '{password}'")

        if not username or not password:
            print("❌ Missing username or password")
            return jsonify({'success': False, 'message': 'Username and password are required'}), 400

        print(f"🔍 Querying Supabase for admin: {username}")
        result = supabase.table('admins').select('*').eq('username', username).execute()
        
        print(f"📊 Query executed")
        print(f"📊 Result type: {type(result)}")
        print(f"📊 Result data: {result.data if hasattr(result, 'data') else 'NO DATA ATTRIBUTE'}")
        
        if result.data and len(result.data) > 0:
            admin = result.data[0]
            print(f"✅ Admin found: {admin}")
            
            # Compare plain password with stored password_hash (both stored as plain text in DB)
            stored_password = admin.get('password_hash', '')
            if stored_password == password:
                print("✅ Password match - Login successful!")
                return jsonify({
                    'success': True,
                    'user': {
                        'username': admin.get('username'),
                        'name': admin.get('name', '')
                    }
                }), 200
            else:
                print(f"❌ Password mismatch!")
                print(f"   Stored: '{stored_password}'")
                print(f"   Provided: '{password}'")
                return jsonify({'success': False, 'message': 'Invalid username or password'}), 401
        else:
            print("❌ No admin found with that username")
            
        return jsonify({'success': False, 'message': 'Invalid username or password'}), 401
        
    except Exception as e:
        print(f"❌ EXCEPTION: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Server error: {str(e)}'}), 500
    finally:
        print("="*80 + "\n")
    
@app.route('/api/admin/change-password', methods=['POST'])
def admin_change_password():
    """Admin change password. Expects JSON: { username, current_password, new_password }"""
    try:
        data = request.get_json() or {}
        username = (data.get('username') or '').strip()
        current_password = (data.get('current_password') or '').strip()
        new_password = (data.get('new_password') or '').strip()

        if not username or not current_password or not new_password:
            return jsonify({'success': False, 'message': 'Username, current password, and new password are required'}), 400

        if len(new_password) < 6:
            return jsonify({'success': False, 'message': 'Password must be at least 6 characters long'}), 400

        # Find admin user
        try:
            result = supabase.table('admins').select('*').eq('username', username).execute()
            if not result.data or len(result.data) == 0:
                return jsonify({'success': False, 'message': 'Admin user not found'}), 404

            admin = result.data[0]
            
            # Check if password_hash column exists and has a value
            if 'password_hash' not in admin or not admin['password_hash']:
                # For new admins without password set, allow setting one without verification
                stored_password = admin.get('password') or admin.get('password_hash')
            else:
                stored_password = admin.get('password_hash')
            
            # Verify current password (in production, use proper password hashing)
            if stored_password and stored_password != current_password:
                return jsonify({'success': False, 'message': 'Current password is incorrect'}), 401
            
            # Update password - try password_hash first, then password
            try:
                # First, try to update with password_hash column
                try:
                    update_result = supabase.table('admins').update({'password_hash': new_password}).eq('username', username).execute()
                    
                    if update_result and update_result.data:
                        return jsonify({'success': True, 'message': 'Password changed successfully'}), 200
                except Exception as hash_error:
                    # If password_hash fails, try password column instead
                    logger.warning(f"Admin password_hash update failed, trying password column: {str(hash_error)}")
                    try:
                        update_result = supabase.table('admins').update({'password': new_password}).eq('username', username).execute()
                        
                        if update_result and update_result.data:
                            return jsonify({'success': True, 'message': 'Password changed successfully'}), 200
                    except Exception as password_error:
                        logger.exception(f"Admin password update failed: {str(password_error)}")
                        raise password_error
                
                # If we get here, something went wrong
                return jsonify({
                    'success': False,
                    'message': 'Failed to update password in database'
                }), 500
            except Exception as update_error:
                error_msg = str(update_error)
                logger.exception(f"Admin password update error: {error_msg}")
                return jsonify({'success': False, 'message': f'Database error: {error_msg}'}), 500
        except Exception as e:
            logger.exception(f"Admin change password error: {str(e)}")
            return jsonify({'success': False, 'message': f'Error changing admin password: {str(e)}'}), 500
    except Exception as e:
        logger.exception(f"Admin change password outer error: {str(e)}")
        return jsonify({'success': False, 'message': f'Admin change password error: {str(e)}'}), 500


@app.route('/api/upload-log', methods=['GET'])
def get_upload_log():
    """Return last N lines from the uploads.log file for debugging."""
    try:
        lines = int(request.args.get('lines', 200))
    except:
        lines = 200

    try:
        if not os.path.exists(LOG_FILE):
            return jsonify({'error': 'Log file not found', 'lines': []}), 404

        with open(LOG_FILE, 'r', encoding='utf-8') as f:
            all_lines = f.readlines()
            last_lines = all_lines[-lines:] if len(all_lines) > lines else all_lines
            return jsonify({'lines': last_lines, 'total': len(all_lines)}), 200
    except Exception as e:
        logger.exception(f"Error reading log file: {str(e)}")
        return jsonify({'error': f'Error reading log file: {str(e)}'}), 500


@app.route('/api/admin/upload-errors', methods=['GET'])
def get_admin_upload_errors():
    """Return structured upload errors for admin dashboard."""
    try:
        limit = int(request.args.get('limit', 50))
    except:
        limit = 50

    try:
        if not os.path.exists(LOG_FILE):
            return jsonify({'errors': [], 'total': 0}), 200

        errors = []
        with open(LOG_FILE, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                
                # Parse lines that contain errors
                if 'ERROR' in line or 'error' in line or 'Error' in line:
                    # Extract timestamp and message
                    try:
                        # Format: 2025-12-21 01:25:26,123 - ERROR - message
                        parts = line.split(' - ', 2)
                        if len(parts) >= 3:
                            timestamp = parts[0]
                            level = parts[1]
                            message = parts[2]
                            errors.append({
                                'timestamp': timestamp,
                                'level': level,
                                'message': message
                            })
                        else:
                            errors.append({
                                'timestamp': '',
                                'level': 'ERROR',
                                'message': line
                            })
                    except Exception as e:
                        errors.append({
                            'timestamp': '',
                            'level': 'ERROR',
                            'message': line
                        })
        
        # Return most recent errors first
        errors_recent = errors[-limit:] if len(errors) > limit else errors
        errors_recent.reverse()
        
        return jsonify({
            'errors': errors_recent,
            'total': len(errors),
            'limit': limit
        }), 200
    except Exception as e:
        logger.exception(f"Error reading upload errors: {str(e)}")
        return jsonify({'error': f'Error reading upload errors: {str(e)}', 'errors': []}), 500



@app.route('/api/export-upload-errors', methods=['GET'])
def export_upload_errors():
    """Scan recent upload logs for error lines and export them to CSV.
    Returns JSON with `file` set to the relative path of the exported CSV.
    """
    try:
        # Ensure exports directory
        exports_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'exports')
        os.makedirs(exports_dir, exist_ok=True)

        # Read last part of the log to avoid huge reads
        with open(LOG_FILE, 'r', encoding='utf-8') as f:
            all_lines = f.readlines()

        # Keep last 5000 lines for scanning
        recent = all_lines[-5000:]

        # Patterns to capture: Missing parent_no, Insert error, Row ...: messages
        import re
        patterns = [r"Missing parent_no", r"Insert error", r"Row .*?:"]
        matches = []
        for line in recent:
            if any(re.search(p, line) for p in patterns):
                # Try to split timestamp and message
                parts = line.strip().split(' - ', 2)
                if len(parts) >= 3:
                    ts, level, msg = parts[0], parts[1], parts[2]
                elif len(parts) == 2:
                    ts, level = parts[0], parts[1]
                    msg = ''
                else:
                    ts = ''
                    level = ''
                    msg = line.strip()
                matches.append({'timestamp': ts, 'level': level, 'message': msg})

        if not matches:
            return jsonify({'error': 'No recent upload errors found'}), 404

        # Write CSV
        import csv
        timestamp = datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')
        csv_name = f'upload_errors_{timestamp}.csv'
        csv_path = os.path.join(exports_dir, csv_name)
        with open(csv_path, 'w', encoding='utf-8', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=['timestamp', 'level', 'message'])
            writer.writeheader()
            for m in matches:
                writer.writerow(m)

        # Return relative path for download by front-end (served statically under uploads/)
        rel_path = f'uploads/exports/{csv_name}'
        logger.info('Exported upload errors to %s', csv_path)
        return jsonify({'file': rel_path, 'count': len(matches)}), 200
    except Exception as e:
        logger.exception('Error exporting upload errors: %s', str(e))
        return jsonify({'error': str(e)}), 500
        with open(LOG_FILE, 'r', encoding='utf-8', errors='ignore') as f:
            last = list(deque(f, maxlen=lines))

        return jsonify({'lines': last, 'count': len(last)}), 200
    except Exception as e:
        logger.exception(f"Error reading log file: {str(e)}")
        return jsonify({'error': f'Error reading log file: {str(e)}'}), 500


@app.route('/api/upload-log/download', methods=['GET'])
def download_upload_log():
    """Serve the uploads.log file as a downloadable attachment."""
    try:
        if not os.path.exists(LOG_FILE):
            return jsonify({'error': 'Log file not found'}), 404

        # Use send_file to return as attachment
        return send_file(LOG_FILE, as_attachment=True, download_name=os.path.basename(LOG_FILE), mimetype='text/plain')
    except Exception as e:
        logger.exception(f"Error sending log file: {str(e)}")
        return jsonify({'error': f'Error sending log file: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)

